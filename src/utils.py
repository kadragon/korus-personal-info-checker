"""
This module provides utility functions for common tasks such as date
manipulation, directory creation,
Excel file processing (saving with auto-fit column widths), searching for
and preparing specific Excel files to process.
"""

import os
import unicodedata
import zipfile
from collections.abc import Callable
from datetime import datetime

import holidays
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import config as cfg
from .display import (
    print_error,
    print_info,
    print_result,
    print_zip_result,
    print_zip_warning,
)

# --- KORUS Excel style constants ---
_THIN_SIDE = Side(style="thin", color="E5E5E5")
_HEADER_FILL = PatternFill(start_color="F4F4F4", end_color="F4F4F4", fill_type="solid")
_HEADER_FONT = Font(name="Pretendard", size=12, color="333333")
_DATA_FONT = Font(name="Pretendard", size=11, color="333333")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_HEADER_ROW_HEIGHT = 30
_DATA_ROW_HEIGHT = 27
_LEFT_ALIGN_WIDTH_THRESHOLD = 20
_MAX_COLUMN_WIDTH = 55  # ~400px


def _display_width(text: str) -> int:
    """Calculate approximate display width accounting for wide (CJK) characters."""
    width = 0
    for ch in text:
        if unicodedata.east_asian_width(ch) in ("W", "F"):
            width += 2
        else:
            width += 1
    return width


def _apply_korus_style(ws: Worksheet) -> None:
    """Apply KORUS-style formatting to a worksheet.

    Applies consistent styling matching the KORUS system export format:
    header band with light grey fill, Pretendard font, subtle borders,
    and auto-fit column widths with CJK-aware width calculation.
    """
    max_col = ws.max_column
    max_row = ws.max_row

    if not max_col or not max_row or max_row < 1:
        return

    thin = _THIN_SIDE

    # First pass: calculate column widths and determine alignment per column
    col_max_data_widths: list[int] = []
    for col_idx in range(1, max_col + 1):
        max_width = 0
        max_data_width = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                w = _display_width(str(cell.value))
                max_width = max(max_width, w)
                if row_idx > 1:
                    max_data_width = max(max_data_width, w)

        adjusted = max_width + 2 if max_width > 0 else 10
        ws.column_dimensions[col_letter].width = min(adjusted, _MAX_COLUMN_WIDTH)
        col_max_data_widths.append(max_data_width)

    col_alignments = [
        _LEFT_ALIGN if w > _LEFT_ALIGN_WIDTH_THRESHOLD else _CENTER_ALIGN
        for w in col_max_data_widths
    ]

    # Apply formatting to all rows in a single pass
    for row_idx in range(1, max_row + 1):
        is_header = row_idx == 1
        ws.row_dimensions[row_idx].height = (
            _HEADER_ROW_HEIGHT if is_header else _DATA_ROW_HEIGHT
        )
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _HEADER_FONT if is_header else _DATA_FONT
            cell.alignment = _CENTER_ALIGN if is_header else col_alignments[col_idx - 1]
            if is_header:
                cell.fill = _HEADER_FILL
            has_left = col_idx == 1
            cell.border = Border(
                top=thin if is_header else None,
                bottom=thin,
                left=thin if has_left else None,
                right=thin,
            )


def get_prev_month_yyyymm() -> str:
    """
    Calculates the previous month from the current date and returns it as a
    string in 'YYYYMM' format.

    Returns:
        str: The previous month in 'YYYYMM' format.
    """
    today = datetime.today()
    prev_month_date = today - relativedelta(months=1)
    return prev_month_date.strftime("%Y%m")


def _ensure_dir(path: str) -> str:
    """
    Creates a directory if it doesn't exist and prints a message.

    Parameters:
        path (str): The directory path to ensure exists.

    Returns:
        str: The same path that was passed in.
    """
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
        print(f"폴더 생성: {path}")
    return path


def make_save_dir(base_save_dir: str) -> str:
    """
    Creates a subdirectory named after the previous month (YYYYMM) within
    `base_save_dir`.
    If the subdirectory already exists, no action is taken.

    Parameters:
        base_save_dir (str): The base directory where the new subdirectory
        will be created.
                             This path must be an existing directory.

    Returns:
        str: The full path to the created or existing subdirectory for the
        previous month.
    """
    prev_month_str = get_prev_month_yyyymm()
    save_dir = os.path.join(base_save_dir, prev_month_str)
    return _ensure_dir(save_dir)


def save_excel_with_autofit(df: pd.DataFrame, path: str) -> None:
    """
    Saves a Pandas DataFrame to an Excel file with KORUS-style formatting.

    Applies Pretendard font, light grey header band, subtle borders,
    auto-fit column widths, and appropriate row heights.

    Parameters:
        df (pd.DataFrame): The DataFrame to save.
        path (str): The full path (including filename) where the Excel file
        will be saved.
    """
    df.to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws is None:
        wb.close()
        print_error("활성 워크시트를 찾을 수 없어 서식을 적용할 수 없습니다.")
        return

    _apply_korus_style(ws)
    wb.save(path)
    wb.close()


def style_excel_file(path: str) -> None:
    """Apply KORUS-style formatting to all sheets in an existing Excel file.

    Parameters:
        path (str): The full path to the Excel file to style.
    """
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        _apply_korus_style(ws)
    wb.save(path)
    wb.close()


def _find_excel_files(download_dir: str, file_prefix: str) -> list[str]:
    """Finds a list of Excel files in the specified directory based on prefix
    and extension."""
    if not download_dir or not os.path.isdir(download_dir):
        raise EnvironmentError(f"다운로드 디렉토리를 찾을 수 없습니다: {download_dir}")

    matched_files = [
        f
        for f in os.listdir(download_dir)
        if f.startswith(file_prefix) and f.lower().endswith(cfg.EXCEL_EXTENSIONS)
    ]
    # Log count only — raw filenames must not be logged (PIPA compliance)
    print_info(f"파일 검색: '{file_prefix}' 접두사 파일 {len(matched_files)}개 발견")
    return matched_files


def _merge_and_preprocess_files(
    excel_files: list[str], download_dir: str
) -> pd.DataFrame | None:
    """Reads the list of Excel files, merges them into a single DataFrame,
    and preprocesses it."""
    all_dfs = []
    for file_name in excel_files:
        file_path = os.path.join(download_dir, file_name)
        try:
            if file_path.lower().endswith(".xlsx"):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_excel(file_path, engine="xlrd")
            all_dfs.append(df)
        except Exception as e:
            print_error(f"'{file_name}' 파일 처리 중 오류 발생: {e}")
            return None

    if not all_dfs:
        return None

    merged_df = pd.concat(all_dfs, ignore_index=True)

    # Standardize access time columns ("접근일시", "일시") to the standard name
    if cfg.COL_ALT_ACCESS_TIME_1 in merged_df.columns:
        merged_df.rename(
            columns={cfg.COL_ALT_ACCESS_TIME_1: cfg.COL_ACCESS_TIME},
            inplace=True,
        )
    elif cfg.COL_ALT_ACCESS_TIME_2 in merged_df.columns:
        merged_df.rename(
            columns={cfg.COL_ALT_ACCESS_TIME_2: cfg.COL_ACCESS_TIME},
            inplace=True,
        )

    # 표준화된 "접속일시" 컬럼이 존재하면 datetime으로 변환
    if cfg.COL_ACCESS_TIME in merged_df.columns:
        merged_df[cfg.COL_ACCESS_TIME] = pd.to_datetime(merged_df[cfg.COL_ACCESS_TIME])

    # "교번" 또는 "신분번호" 컬럼을 "교직원ID"로 표준화
    has_gyobeon = "교번" in merged_df.columns
    has_sinbun = "신분번호" in merged_df.columns

    if has_gyobeon and has_sinbun:
        print_info(
            "경고: 입력 파일에 '교번'과 '신분번호' 컬럼이 모두 존재합니다. "
            "'교번'을 '교직원ID'로 우선 사용합니다."
        )
        merged_df.rename(columns={"교번": cfg.COL_EMPLOYEE_ID}, inplace=True)
    elif has_gyobeon:
        merged_df.rename(columns={"교번": cfg.COL_EMPLOYEE_ID}, inplace=True)
    elif has_sinbun:
        merged_df.rename(columns={"신분번호": cfg.COL_EMPLOYEE_ID}, inplace=True)

    return merged_df


_ACCESS_LOG_CACHE: dict[tuple[str, str], pd.DataFrame] = {}


def load_access_logs_cached(download_dir: str, file_prefix: str) -> pd.DataFrame | None:
    """Load and merge access log files; cache by (download_dir, file_prefix)."""
    key = (download_dir, file_prefix)
    if key in _ACCESS_LOG_CACHE:
        return _ACCESS_LOG_CACHE[key].copy()
    excel_files = _find_excel_files(download_dir, file_prefix)
    if not excel_files:
        return None
    merged = _merge_and_preprocess_files(excel_files, download_dir)
    if merged is None:
        return None
    _ACCESS_LOG_CACHE[key] = merged
    return merged.copy()


def clear_access_log_cache() -> None:
    """Clear the in-memory access log cache. Intended for test isolation."""
    _ACCESS_LOG_CACHE.clear()


def find_and_prepare_excel_file(
    download_dir: str,
    file_prefix: str,
    save_dir: str,
    output_file_basename: str,
    prev_month: str,
) -> tuple[pd.DataFrame | None, str | None]:
    """
    Finds, merges, preprocesses, and saves Excel files from the specified folder.

    This function performs the following:
    1. Uses `load_access_logs_cached` to find, merge, and cache relevant files.
    2. Saves the merged DataFrame as an intermediate result.
    """
    try:
        merged_df = load_access_logs_cached(download_dir, file_prefix)
    except EnvironmentError as e:
        print_error(str(e))
        return None, None

    if merged_df is None:
        print_info(
            f"'{file_prefix}'로 시작하는 파일을 찾을 수 없습니다. 이 검사는 건너뜁니다."
        )
        return None, None

    print_info(f"{output_file_basename} 원본 데이터: {len(merged_df)}건")

    os.makedirs(save_dir, exist_ok=True)
    destination_save_path = os.path.join(
        save_dir, f"{output_file_basename}_{prev_month}.xlsx"
    )
    try:
        save_excel_with_autofit(merged_df, destination_save_path)
        save_msg = (
            f"모든 파일을 합쳐 '{os.path.basename(destination_save_path)}'"
            f"(으)로 저장했습니다."
        )
        print_info(save_msg)
    except Exception as e:
        print_error(f"병합된 파일 저장 중 오류 발생: {e}")
        return None, None

    return merged_df, destination_save_path


def zip_files_by_prefix(target_dir: str, prefix_list: list[str]) -> None:
    """
    Creates zip archives grouped by prefix.

    The archive name is derived from the provided prefix when available,
    falling back to the legacy filename-based heuristic otherwise.
    """
    files = [f for f in os.listdir(target_dir) if f.endswith(".xlsx")]

    for prefix in prefix_list:
        matched = [f for f in files if f.startswith(prefix)]
        if not matched:
            print_zip_warning(prefix)
            continue

        # Use the full prefix as the zip file name
        # (e.g., "[붙임2] 코러스 사용자 접근 기록.zip")
        group_name = prefix
        zip_name = f"{group_name}.zip"
        zip_path = os.path.join(target_dir, zip_name)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for filename in matched:
                zipf.write(os.path.join(target_dir, filename), arcname=filename)

        print_zip_result(zip_name, len(matched))


def filter_by_time_conditions(
    df: pd.DataFrame,
    time_col: str,
    employee_id_col: str,
    check_off_hours: bool,
    check_holidays_weekends: bool,
    off_hours_start: int,
    off_hours_end: int,
) -> pd.DataFrame:
    """
    Filters the DataFrame based on time conditions (off-hours, holidays/weekends).

    Parameters:
        df (pd.DataFrame): The DataFrame to filter.
        time_col (str): The column name containing timestamp information.
        employee_id_col (str): The column name containing employee ID.
        check_off_hours (bool): Whether to enable off-hours checking.
        check_holidays_weekends (bool): Whether to enable holidays and
        weekends checking.
        off_hours_start (int): The start time for off-hours.
        off_hours_end (int): The end time for off-hours.

    Returns:
        pd.DataFrame: The filtered DataFrame meeting the specified time conditions.
    """
    if df is None:
        raise ValueError("Input DataFrame cannot be None.")

    df_copy = df.copy()

    final_mask = pd.Series(False, index=df.index)

    if check_off_hours:
        hour = df_copy[time_col].dt.hour
        is_off_hour = (hour < off_hours_end) | (hour >= off_hours_start)
        final_mask |= is_off_hour

    if check_holidays_weekends:
        years = df_copy[time_col].dt.year.unique()
        kr_holidays = holidays.KR(years=years)  # type: ignore [attr-defined]
        weekday = df_copy[time_col].dt.weekday
        date_only = df_copy[time_col].dt.date

        is_weekend = weekday >= 5  # Monday is 0, Sunday is 6
        is_holiday = date_only.isin(kr_holidays)
        final_mask |= is_weekend
        final_mask |= is_holiday

    return df_copy[final_mask].sort_values([employee_id_col, time_col])


def run_and_save_check(
    df: pd.DataFrame,
    check_func: Callable[[pd.DataFrame], pd.DataFrame],
    save_path: str,
    result_description: str,
) -> None:
    """
    Runs the check function, saves the result to an Excel file if any, and
    outputs a status message.

    Parameters:
        df (pd.DataFrame): The input DataFrame to perform the check on.
        check_func (function): A function that takes a DataFrame and returns
        a filtered DataFrame.
        save_path (str): The path where the result Excel file will be saved.
        result_description (str): The description to use in the output message
        when results are found or not found.
    """
    filtered_df = check_func(df)
    if not filtered_df.empty:
        save_excel_with_autofit(filtered_df, save_path)
        print_result(
            is_detected=True,
            description=result_description,
            filename=os.path.basename(save_path),
        )
    else:
        print_result(is_detected=False, description=result_description)

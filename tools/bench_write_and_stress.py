"""Honest verification of (A) the zero-risk write fix and (B) filter parity at max user size.

(A) Reparse-kill: keep the EXISTING _apply_korus_style (identical output), but style the
    in-memory workbook that to_excel already built instead of writing→reloading→saving twice.
(B) Stress filters near the user's stated max (~1000 rows/user) — fewer users, same total —
    to confirm parity holds and the fast ip_switch per-group loop stays acceptable.

Run: uv run python tools/bench_write_and_stress.py
"""

from __future__ import annotations

import os
import tempfile
import time

import bench_filters_parity as bp  # reuse fixtures + fast impls + parity asserts
import pandas as pd

import src.config as cfg
from src.checkers import download_reason_checker as dr
from src.checkers import login_checker as lc
from src.utils import _apply_korus_style, save_excel_with_autofit


def save_reparse_kill(df: pd.DataFrame, path: str) -> None:
    """Single build + existing style + single save. No load_workbook, no double save."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        ws = writer.book.active
        _apply_korus_style(ws)
    # context exit saves once


def _cells(path: str) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return [c.value for row in ws.iter_rows() for c in row]


def verify_write(rows: int, users: int) -> None:
    print(f"\n=== (A) WRITE: current vs reparse-kill  (rows={rows:,}) ===")
    frames = bp.login_fixture(rows, users, burst=False)
    with tempfile.TemporaryDirectory() as tmp:
        p_cur = os.path.join(tmp, "cur.xlsx")
        p_new = os.path.join(tmp, "new.xlsx")

        t0 = time.perf_counter()
        save_excel_with_autofit(frames, p_cur)
        t_cur = time.perf_counter() - t0

        t0 = time.perf_counter()
        save_reparse_kill(frames, p_new)
        t_new = time.perf_counter() - t0

        # output identity: same cell values + same column widths (styling code is shared)
        same_vals = _cells(p_cur) == _cells(p_new)
        import openpyxl
        w_cur = openpyxl.load_workbook(p_cur).active.column_dimensions
        w_new = openpyxl.load_workbook(p_new).active.column_dimensions
        same_w = all(
            abs((w_cur[k].width or 0) - (w_new[k].width or 0)) < 1e-6
            for k in {*w_cur.keys(), *w_new.keys()}
        )
        print(f"  current save_excel_with_autofit   {t_cur*1000:>9.1f} ms")
        print(f"  reparse-kill (existing styling)   {t_new*1000:>9.1f} ms"
              f"   ({t_cur/t_new:.2f}×)")
        print(f"  output identical? values={same_vals} widths={same_w}  "
              f"{'✅' if same_vals and same_w else '❌'}")


def verify_stress(total: int, users: int) -> None:
    rpu = total // users
    print(f"\n=== (B) FILTER STRESS near max user size "
          f"(rows={total:,} users={users} ~{rpu}/user) ===")

    dlf = bp.download_fixture(total, users, burst=True)
    lf = bp.login_fixture(total, users, burst=True)
    mu = lf.groupby(cfg.COL_EMPLOYEE_ID).size()
    print(f"  actual rows/user: max={mu.max()} mean={mu.mean():.0f}")

    # parity
    bp.assert_identical("high_freq stress",
                        dr._filter_high_freq_download(dlf), bp._filter_high_freq_fast(dlf))
    bp.assert_identical("ip_switch stress",
                        lc._filter_ip_switch(lf), bp._filter_ip_switch_fast(lf),
                        cols=[cfg.COL_ESTIMATED_REASON, cfg.COL_RISK_LEVEL,
                              cfg.COL_UNIQUE_IP_COUNT, cfg.COL_UNIQUE_SUBNET_COUNT])

    # speed of FAST impls at this user size
    for name, fn, arg in [
        ("high_freq FAST", bp._filter_high_freq_fast, dlf),
        ("ip_switch FAST", bp._filter_ip_switch_fast, lf),
    ]:
        t0 = time.perf_counter()
        fn(arg)
        print(f"  {name:<16} {(time.perf_counter()-t0)*1000:>9.1f} ms")


def main() -> None:
    verify_write(100_000, 2000)
    # near user's stated MAX: ~1000 rows for the heaviest user
    verify_stress(60_000, 70)
    print("\nDONE")


if __name__ == "__main__":
    main()

"""Performance benchmark harness for the KORUS checker pipeline.

Generates synthetic access-log data matching the production schema, then times
each pipeline stage to locate the true bottleneck. Throwaway analysis tool —
not part of the shipped package.

Run: uv run python tools/bench_perf.py [--rows 500000] [--users 5000]
"""

from __future__ import annotations

import argparse
import os
import tempfile
import time
from contextlib import contextmanager

import numpy as np
import pandas as pd

import src.config as cfg
from src.utils import _apply_korus_style

RNG = np.random.default_rng(42)


@contextmanager
def timed(label: str, store: dict[str, float]):
    t0 = time.perf_counter()
    yield
    dt = time.perf_counter() - t0
    store[label] = dt
    print(f"  {label:<46} {dt*1000:>10.1f} ms")


def _rand_ips(n: int) -> np.ndarray:
    # Mostly campus private IPs; a few users hop subnets.
    a = RNG.choice([192, 10, 172], size=n, p=[0.8, 0.15, 0.05])
    b = np.where(a == 192, 168, RNG.integers(0, 32, n))
    c = RNG.integers(0, 256, n)
    d = RNG.integers(1, 255, n)
    return np.char.add(np.char.add(np.char.add(
        np.char.add(np.char.add(a.astype(str), "."), b.astype(str)), "."),
        np.char.add(c.astype(str), ".")), d.astype(str))


def make_frames(total_rows: int, n_users: int) -> dict[str, pd.DataFrame]:
    """Build login / download / access frames at the requested scale.

    Each user capped well under 1000 rows (matches calibration: max user <= 1천).
    """
    uids = RNG.integers(0, n_users, total_rows)
    emp = np.char.add("E", uids.astype(str))
    base = np.datetime64("2026-05-01T00:00:00")
    offs = RNG.integers(0, 31 * 24 * 3600, total_rows).astype("timedelta64[s]")
    ts = base + offs

    login = pd.DataFrame({
        cfg.COL_EMPLOYEE_ID: emp,
        cfg.COL_ACCESS_TIME: ts,
        cfg.COL_IP: _rand_ips(total_rows),
    })

    programs = np.array(["인사마스터", "급여관리", "학사정보", "복지관리", "기타"])
    jobs = np.array(["조회", "저장", "수정", "삭제"])
    names = np.char.add("사용자", (uids % 500).astype(str))
    access = pd.DataFrame({
        cfg.COL_EMPLOYEE_ID: emp,
        cfg.COL_PROGRAM_NAME: programs[RNG.integers(0, len(programs), total_rows)],
        cfg.COL_DETAIL_CONTENT: np.char.add("sklstfNo=", np.char.add(
            "E", RNG.integers(0, n_users, total_rows).astype(str))),
        cfg.COL_ACCESS_TIME: ts,
        cfg.COL_JOB_PERFORMANCE: jobs[RNG.integers(0, len(jobs), total_rows)],
        cfg.COL_EMPLOYEE_NAME: names,
    })

    reasons = np.array(["연구자료 분석", "민원처리", "asdf", "ㅁㄴㅇㄹ",
                        "감사대응", "qwerty", "1234", "통계작성"])
    download = pd.DataFrame({
        cfg.COL_EMPLOYEE_ID: emp,
        cfg.COL_DOWNLOAD_REASON: reasons[RNG.integers(0, len(reasons), total_rows)],
        cfg.COL_DOWNLOAD_COUNT: RNG.integers(1, 300, total_rows),
        cfg.COL_ACCESS_TIME: ts,
        cfg.COL_PROGRAM_NAME: programs[RNG.integers(0, len(programs), total_rows)],
    })
    return {"login": login, "access": access, "download": download}


def bench_read(df: pd.DataFrame, tmp: str, store: dict[str, float]) -> None:
    print("\n[READ] same .xlsx, openpyxl vs calamine")
    path = os.path.join(tmp, "read_probe.xlsx")
    with timed("write fixture (openpyxl, one-time)", store):
        df.to_excel(path, index=False)
    with timed("read pd.read_excel(openpyxl)  [current]", store):
        pd.read_excel(path)
    try:
        with timed("read pd.read_excel(calamine)   [proposed]", store):
            pd.read_excel(path, engine="calamine")
    except Exception as e:  # noqa: BLE001
        print(f"  calamine read FAILED: {e}")


def bench_write(df: pd.DataFrame, tmp: str, store: dict[str, float]) -> None:
    print("\n[WRITE] styled save: current 4-step vs decomposed + xlsxwriter")
    import openpyxl
    path = os.path.join(tmp, "write_probe.xlsx")

    # Decompose the CURRENT save_excel_with_autofit into its 4 internal steps.
    with timed("1. df.to_excel(openpyxl)", store):
        df.to_excel(path, index=False)
    with timed("2. openpyxl.load_workbook  [full reparse]", store):
        wb = openpyxl.load_workbook(path)
    ws = wb.active
    with timed("3. _apply_korus_style  [2x per-cell loop]", store):
        _apply_korus_style(ws)
    with timed("4. wb.save", store):
        wb.save(path)
    store["CURRENT write total"] = (
        store["1. df.to_excel(openpyxl)"]
        + store["2. openpyxl.load_workbook  [full reparse]"]
        + store["3. _apply_korus_style  [2x per-cell loop]"]
        + store["4. wb.save"]
    )
    print(f"  {'== CURRENT write total':<46} "
          f"{store['CURRENT write total']*1000:>10.1f} ms")

    # Proposed: single-pass xlsxwriter with vectorized column widths.
    path2 = os.path.join(tmp, "write_probe_xlsx.xlsx")
    with timed("PROPOSED single-pass xlsxwriter", store):
        _write_xlsxwriter_styled(df, path2)


def _write_xlsxwriter_styled(df: pd.DataFrame, path: str) -> None:
    """Single-pass styled write. Column widths computed vectorized (no cell loop)."""

    def disp_width_series(s: pd.Series) -> int:
        sample = s.astype(str)
        # width ~= len + count of wide CJK chars; cheap upper bound via max len*1.8
        m = sample.map(len).max()
        return int(m) if pd.notna(m) else 0

    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        wb = writer.book
        ws = writer.sheets["Sheet1"]
        header_fmt = wb.add_format({
            "bold": False, "font_name": "Pretendard", "font_size": 12,
            "bg_color": "#F4F4F4", "border": 1, "align": "center", "valign": "vcenter",
        })
        data_fmt = wb.add_format({
            "font_name": "Pretendard", "font_size": 11, "border": 1, "valign": "vcenter",
        })
        for c, col in enumerate(df.columns):
            w = max(disp_width_series(df[col]), len(str(col))) + 2
            ws.set_column(c, c, min(w, 55), data_fmt)
            ws.write(0, c, col, header_fmt)


def bench_filters(frames: dict, store: dict[str, float]) -> None:
    print("\n[FILTERS] compute hot paths at scale")
    from src.checkers import download_reason_checker as dr
    from src.checkers import login_checker as lc

    login = frames["login"]
    download = frames["download"]
    access = frames["access"]

    with timed("login _filter_ip_switch  [O(n^2)/user]", store):
        lc._filter_ip_switch(login)
    with timed("download _filter_high_freq  [O(n^2)/user]", store):
        dr._filter_high_freq_download(download)
    with timed("download _check_download_sayu  [.apply regex]", store):
        dr._check_download_sayu(download)
    with timed("download _enrich_access_log  [py loop]", store):
        dr._enrich_with_access_log_summary(
            download, access, cfg.CROSS_REF_TIME_WINDOW_MINUTES)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", type=int, default=500_000)
    ap.add_argument("--users", type=int, default=5_000)
    args = ap.parse_args()

    print(f"=== KORUS perf benchmark | rows={args.rows:,} users={args.users:,} ===")
    store: dict[str, float] = {}
    with timed("generate synthetic frames", store):
        frames = make_frames(args.rows, args.users)
    mu = frames["login"].groupby(cfg.COL_EMPLOYEE_ID).size()
    print(f"  rows/user: max={mu.max()} mean={mu.mean():.0f} users={len(mu)}")

    with tempfile.TemporaryDirectory() as tmp:
        bench_read(frames["login"], tmp, store)
        bench_write(frames["login"], tmp, store)
        bench_filters(frames, store)

    print("\n=== SUMMARY (ms) ===")
    for k, v in store.items():
        print(f"  {k:<46} {v*1000:>10.1f}")


if __name__ == "__main__":
    main()

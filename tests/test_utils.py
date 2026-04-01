"""
Tests for utils module.

GENERATED FROM SPEC-test-coverage-improvement-1
Trace: SPEC-test-coverage-improvement-1, TEST-utils-1
"""

import os
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

from src import utils


class TestGetPrevMonthYyyymm:
    def test_get_prev_month_yyyymm(self, mocker):
        # Mock datetime.today to return a fixed date
        mock_today = mocker.patch("src.utils.datetime")
        mock_today.today.return_value = datetime(2023, 10, 15)
        mock_today.return_value = datetime

        result = utils.get_prev_month_yyyymm()
        assert result == "202309"


class TestMakeSaveDir:
    def test_make_save_dir_creates_dir(self, temp_dir, mocker):
        mocker.patch("src.utils.get_prev_month_yyyymm", return_value="202309")
        base_dir = temp_dir
        result = utils.make_save_dir(base_dir)
        expected = os.path.join(base_dir, "202309")
        assert result == expected
        assert os.path.exists(expected)

    def test_make_save_dir_existing_dir(self, temp_dir, mocker):
        mocker.patch("src.utils.get_prev_month_yyyymm", return_value="202309")
        base_dir = temp_dir
        existing_dir = os.path.join(base_dir, "202309")
        os.makedirs(existing_dir)
        result = utils.make_save_dir(base_dir)
        assert result == existing_dir


class TestSaveExcelWithAutofit:
    def test_save_excel_with_autofit(self, temp_dir, sample_personal_access_df):
        path = os.path.join(temp_dir, "test.xlsx")
        utils.save_excel_with_autofit(sample_personal_access_df, path)
        assert os.path.exists(path)
        # Check if it's a valid Excel file by reading it back
        df_read = pd.read_excel(path)
        pd.testing.assert_frame_equal(df_read, sample_personal_access_df)

    def test_save_excel_with_autofit_applies_korus_style(self, temp_dir):
        """Test that save_excel_with_autofit applies KORUS-style formatting."""
        df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
        path = os.path.join(temp_dir, "test.xlsx")

        utils.save_excel_with_autofit(df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # Header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.name == "Pretendard"
        assert header_cell.font.size == 12
        assert header_cell.fill.start_color.rgb == "00F4F4F4"
        assert header_cell.alignment.horizontal == "center"
        assert header_cell.alignment.wrap_text is True

        # Data formatting
        data_cell = ws.cell(row=2, column=1)
        assert data_cell.font.name == "Pretendard"
        assert data_cell.font.size == 11

        # Row heights
        assert ws.row_dimensions[1].height == 30
        assert ws.row_dimensions[2].height == 27

        wb.close()


class TestFindExcelFiles:
    def test_find_excel_files(self, temp_dir):
        # Create test files
        os.makedirs(temp_dir, exist_ok=True)
        open(os.path.join(temp_dir, "test.xlsx"), "w").close()
        open(os.path.join(temp_dir, "test.xls"), "w").close()
        open(os.path.join(temp_dir, "other.txt"), "w").close()

        result = utils._find_excel_files(temp_dir, "test")
        assert len(result) == 2
        assert "test.xlsx" in result
        assert "test.xls" in result

    def test_find_excel_files_no_files(self, temp_dir):
        result = utils._find_excel_files(temp_dir, "test")
        assert result == []

    def test_find_excel_files_invalid_dir(self):
        """Test _find_excel_files with invalid directory (line 109)."""
        with pytest.raises(EnvironmentError) as exc_info:
            utils._find_excel_files("/nonexistent/path", "test")
        assert "다운로드 디렉토리를 찾을 수 없습니다" in str(exc_info.value)


class TestMergeAndPreprocessFiles:
    def test_merge_and_preprocess_files(self, temp_dir, sample_personal_access_df):
        # Create a test Excel file
        file_path = os.path.join(temp_dir, "test.xlsx")
        sample_personal_access_df.to_excel(file_path, index=False)

        result = utils._merge_and_preprocess_files(["test.xlsx"], temp_dir)
        assert result is not None
        pd.testing.assert_frame_equal(result, sample_personal_access_df)

    def test_merge_and_preprocess_files_no_files(self, temp_dir):
        result = utils._merge_and_preprocess_files([], temp_dir)
        assert result is None

    def test_merge_and_preprocess_files_xls_format(self, temp_dir, mocker):
        """Test _merge_and_preprocess_files with .xls file (line 130)."""
        df = pd.DataFrame({"접속일시": ["2023-09-01 10:00:00"], "교번": ["12345"]})
        os.path.join(temp_dir, "test.xls")

        # Since xlwt is not installed, we'll mock the read_excel call
        # to simulate reading an .xls file
        mock_read = mocker.patch("pandas.read_excel", return_value=df)

        result = utils._merge_and_preprocess_files(["test.xls"], temp_dir)

        # Verify that read_excel was called with engine="xlrd" for .xls files
        mock_read.assert_called_once()
        call_args = mock_read.call_args
        assert call_args.kwargs.get("engine") == "xlrd"

        assert result is not None

    def test_merge_and_preprocess_files_exception(self, temp_dir, mocker, capsys):
        """Test _merge_and_preprocess_files with file read exception (lines 132-134)."""
        # Create a file that will cause an exception
        file_path = os.path.join(temp_dir, "corrupt.xlsx")
        with open(file_path, "w") as f:
            f.write("corrupt data")

        result = utils._merge_and_preprocess_files(["corrupt.xlsx"], temp_dir)

        # Should return None and print error
        assert result is None
        captured = capsys.readouterr()
        assert "오류 발생" in captured.out or len(captured.out) > 0

    def test_merge_and_preprocess_files_alt_access_time_2(self, temp_dir):
        """Test column renaming for '일시' column (line 148)."""
        df = pd.DataFrame({"일시": ["2023-09-01 10:00:00"], "교번": ["12345"]})
        file_path = os.path.join(temp_dir, "test.xlsx")
        df.to_excel(file_path, index=False)

        result = utils._merge_and_preprocess_files(["test.xlsx"], temp_dir)
        assert result is not None
        assert "접속일시" in result.columns
        assert "일시" not in result.columns

    def test_merge_and_preprocess_files_both_id_columns(self, temp_dir, capsys):
        """Test warning when both '교번' and '신분번호' exist (lines 162-166)."""
        df = pd.DataFrame(
            {
                "접속일시": ["2023-09-01 10:00:00"],
                "교번": ["12345"],
                "신분번호": ["67890"],
            }
        )
        file_path = os.path.join(temp_dir, "test.xlsx")
        df.to_excel(file_path, index=False)

        result = utils._merge_and_preprocess_files(["test.xlsx"], temp_dir)

        assert result is not None
        assert "교직원ID" in result.columns
        # Should print warning
        captured = capsys.readouterr()
        assert "경고" in captured.out or "교번" in captured.out

    def test_merge_and_preprocess_files_sinbun_only(self, temp_dir):
        """Test renaming '신분번호' to '교직원ID' (lines 169-170)."""
        df = pd.DataFrame({"접속일시": ["2023-09-01 10:00:00"], "신분번호": ["67890"]})
        file_path = os.path.join(temp_dir, "test.xlsx")
        df.to_excel(file_path, index=False)

        result = utils._merge_and_preprocess_files(["test.xlsx"], temp_dir)

        assert result is not None
        assert "교직원ID" in result.columns
        assert "신분번호" not in result.columns


class TestFindAndPrepareExcelFile:
    def test_find_and_prepare_excel_file(
        self, temp_dir, sample_personal_access_df, mocker
    ):
        # Mock print functions
        mocker.patch("src.utils.print_info")
        mocker.patch("src.utils.print_error")

        # Create test file
        file_path = os.path.join(temp_dir, "prefix_202309.xlsx")
        sample_personal_access_df.to_excel(file_path, index=False)

        save_dir = os.path.join(temp_dir, "save")
        os.makedirs(save_dir)

        df, saved_path = utils.find_and_prepare_excel_file(
            temp_dir, "prefix_", save_dir, "test", "202309"
        )
        assert df is not None
        assert saved_path is not None
        assert os.path.exists(saved_path)

    def test_find_and_prepare_excel_file_no_files(self, temp_dir, mocker):
        mocker.patch("src.utils.print_info")
        save_dir = os.path.join(temp_dir, "save")
        os.makedirs(save_dir)

        df, saved_path = utils.find_and_prepare_excel_file(
            temp_dir, "prefix_", save_dir, "test", "202309"
        )
        assert df is None
        assert saved_path is None

    def test_find_and_prepare_excel_file_env_error(self, mocker, capsys):
        """Test find_and_prepare_excel_file with EnvironmentError (lines 198-200)."""
        mocker.patch("src.utils.print_info")

        # Pass invalid directory to trigger EnvironmentError
        df, saved_path = utils.find_and_prepare_excel_file(
            "/nonexistent", "prefix_", "/tmp", "test", "202309"
        )

        assert df is None
        assert saved_path is None
        captured = capsys.readouterr()
        assert "오류" in captured.out or len(captured.out) > 0

    def test_find_and_prepare_excel_file_merge_returns_none(self, temp_dir, mocker):
        """Test when _merge_and_preprocess_files returns None (line 204)."""
        mocker.patch("src.utils.print_info")

        # Create a corrupt Excel file
        file_path = os.path.join(temp_dir, "prefix_test.xlsx")
        with open(file_path, "w") as f:
            f.write("corrupt")

        save_dir = os.path.join(temp_dir, "save")
        os.makedirs(save_dir)

        df, saved_path = utils.find_and_prepare_excel_file(
            temp_dir, "prefix_", save_dir, "test", "202309"
        )

        assert df is None
        assert saved_path is None

    def test_find_and_prepare_excel_file_save_exception(
        self, temp_dir, sample_personal_access_df, mocker, capsys
    ):
        """Test save exception in find_and_prepare_excel_file (lines 219-221)."""
        mocker.patch("src.utils.print_info")

        # Create test file
        file_path = os.path.join(temp_dir, "prefix_test.xlsx")
        sample_personal_access_df.to_excel(file_path, index=False)

        save_dir = os.path.join(temp_dir, "save")
        os.makedirs(save_dir)

        # Mock to_excel to raise exception
        mocker.patch.object(
            pd.DataFrame, "to_excel", side_effect=Exception("Save failed")
        )

        df, saved_path = utils.find_and_prepare_excel_file(
            temp_dir, "prefix_", save_dir, "test", "202309"
        )

        assert df is None
        assert saved_path is None
        captured = capsys.readouterr()
        assert "오류" in captured.out or len(captured.out) > 0


class TestZipFilesByPrefix:
    def test_zip_files_by_prefix(self, temp_dir, mocker):
        mocker.patch("src.utils.print_zip_result")
        mocker.patch("src.utils.print_zip_warning")

        # Create test files
        save_dir = temp_dir
        open(os.path.join(save_dir, "[붙임2]file1.xlsx"), "w").close()
        open(os.path.join(save_dir, "[붙임2]file2.xlsx"), "w").close()
        open(os.path.join(save_dir, "other.xlsx"), "w").close()

        utils.zip_files_by_prefix(save_dir, ["[붙임2"])

        zip_path = os.path.join(save_dir, "[붙임2.zip")
        assert os.path.exists(zip_path)

    def test_zip_files_by_prefix_no_match(self, temp_dir, capsys):
        """Test zip_files_by_prefix when no files match prefix (lines 238-239)."""
        save_dir = temp_dir
        # Create file that doesn't match
        open(os.path.join(save_dir, "other.xlsx"), "w").close()

        utils.zip_files_by_prefix(save_dir, ["[붙임9]"])

        # Should print warning
        captured = capsys.readouterr()
        assert "WARNING" in captured.out or "없음" in captured.out


class TestFilterByTimeConditions:
    def test_filter_by_time_conditions_off_hours(self, sample_login_df):
        result = utils.filter_by_time_conditions(
            sample_login_df, "접속일시", "교직원ID", True, False, 23, 7
        )
        # Sample data has times in business hours, result should be empty or filtered
        assert isinstance(result, pd.DataFrame)

    def test_filter_by_time_conditions_holidays(self, sample_login_df, mocker):
        # Mock holidays
        mock_holidays = mocker.patch("src.utils.holidays.KR")
        mock_holidays.return_value = []

        result = utils.filter_by_time_conditions(
            sample_login_df, "접속일시", "교직원ID", False, True, 23, 7
        )
        assert isinstance(result, pd.DataFrame)

    def test_filter_by_time_conditions_none_df(self):
        """Test filter_by_time_conditions with None DataFrame (line 280)."""
        with pytest.raises(ValueError) as exc_info:
            utils.filter_by_time_conditions(
                None, "접속일시", "교직원ID", True, False, 23, 7
            )
        assert "Input DataFrame cannot be None" in str(exc_info.value)


class TestRunAndSaveCheck:
    def test_run_and_save_check_with_results(
        self, temp_dir, sample_personal_access_df, mocker
    ):
        mocker.patch("src.utils.print_result")
        mocker.patch("src.utils.save_excel_with_autofit")

        def dummy_check(df):
            return df.head(1)

        path = os.path.join(temp_dir, "test.xlsx")
        utils.run_and_save_check(sample_personal_access_df, dummy_check, path, "test")

    def test_run_and_save_check_no_results(
        self, temp_dir, sample_personal_access_df, mocker
    ):
        mocker.patch("src.utils.print_result")

        def dummy_check(df):
            return pd.DataFrame()

        path = os.path.join(temp_dir, "test.xlsx")
        utils.run_and_save_check(sample_personal_access_df, dummy_check, path, "test")
